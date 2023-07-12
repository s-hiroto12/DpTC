import numpy as np
import pandas as pd
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import BernoulliNB
from sklearn import svm
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import roc_auc_score
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score
from sklearn.metrics import f1_score
import xlwt
import openpyxl
import os
import random
import warnings
warnings.filterwarnings("ignore")

def write_excel_xls(path, sheet_name, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    print("xls格式表格写入数据成功！")

def write_excel_xls_hypoth(path, sheet_name, value):
    if os.path.exists(path):
        data = openpyxl.load_workbook(path)
        table = data.create_sheet(sheet_name)
        table.title = sheet_name
        index = len(value)  # 获取需要写入数据的行数
        for i in range(0, index):
            for j in range(0, len(value[i])):
                table.cell(row=1 + i, column=j + 1, value=str(value[i][j]))  # 像表格中写入数据（对应的行和列）
        data.save(path)  # 保存工作簿
    else:
        index = len(value)
        workbook = openpyxl.Workbook()  # 新建工作簿（默认有一个sheet？）
        sheet = workbook.active  # 获得当前活跃的工作页，默认为第一个工作页
        sheet.title = sheet_name  # 给sheet页的title赋值
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  # 行，列，值 这里是从1开始计数的
        workbook.save(path)  # 一定要保存
    print("xlsx格式表格写入数据成功！")

# ast_whole
def get_data_ast(project):
    root = 'data/ast-whole/' 
    path = root + project + '/features.npy'
    ast_f = np.load(path, allow_pickle=True)
    path_label = root + project + '/labels.npy'
    labels = np.load(path_label, allow_pickle=True)

    print('ast_f sahpe = ',ast_f.shape)

    return ast_f, labels

# 传统特征
def get_data_tradition_10(project):
    path = 'data/tr/' + project +'/'+project+'.npy'
    tr = np.load(path, allow_pickle=True)
    tr_f = tr[:,:-1]
    tr_l = tr[:,-1]
    return tr_f, tr_l  # np.array

# split data for training, and testing, 并进行10折交叉验证
def split_data_train(features, labels, clf):
    labels = labels.reshape((len(labels), 1))
    data = np.hstack((features, labels))
    EPOCHS = 30
    result = []
    result_avg = []
    total_acc = 0.0
    total_auc = 0.0
    total_f1 = 0.0
    total_pre = 0.0
    total_recall = 0.0
    for epoch in range(EPOCHS):
        data_cross = pd.DataFrame(data)
        data_num = len(data_cross)
        k = 10  # 10折交叉验证
        ratio = data_num // k
        data_cross = data_cross.sample(frac=1, random_state= random.randint(1, 1000))

        acc = 0
        auc = 0
        f1 = 0
        pre = 0
        recall = 0
        for i in range(k):
            if i == k - 1:
                test = data_cross.iloc[ratio * i: ratio * (i + 1)]
                train = data_cross.iloc[:ratio * i]
            else:
                test = data_cross.iloc[ratio * i: ratio * (i + 1)]
                train = data_cross.iloc[:ratio * i].append(data_cross.iloc[ratio * (i + 1):])
            train = train.values  # 转换成np.array
            test = test.values
            train_col = train.shape[1]
            X_train = train[:, list(range(train_col - 1))]
            y_train = train[:, -1]
            X_test = test[:, list(range(train_col - 1))]
            y_test = test[:, -1]

            # 开始训练
            clf.fit(X_train, y_train.ravel())
            y_pred = clf.predict(X_test)
            acc += accuracy_score(y_test, y_pred)
            auc += roc_auc_score(y_test, y_pred)
            f1 += f1_score(y_test, y_pred)
            pre += precision_score(y_test, y_pred)
            recall += recall_score(y_test, y_pred)

        result.append([acc / k, auc / k, f1 / k, pre / k, recall / k])

        total_acc += acc / k
        total_auc += auc / k
        total_f1 += f1 / k
        total_pre += pre / k
        total_recall += recall / k
    result_avg = [total_acc/EPOCHS, total_auc/EPOCHS, total_f1/EPOCHS, total_pre/EPOCHS, total_recall/EPOCHS]
    print("Testing results(Acc):", total_acc / EPOCHS)
    print("Testing results(Auc):", total_auc / EPOCHS)
    print("Testing results(F1):", total_f1 / EPOCHS)
    print("Testing results(Pre):", total_pre / EPOCHS)
    print("Testing results(Recall):", total_recall / EPOCHS)
    return result, result_avg


def main():
    projects = ["ambari"]

    logist = LogisticRegression(max_iter=500)   
    nb = BernoulliNB()
    svm_ = svm.SVC()
    randomForest = RandomForestClassifier()  
    models = [logist, randomForest, nb, svm_]

    for j in range(1):
        root = 'result/ast_whole/logit/'
        if not os.path.exists(root):
            os.mkdir(root)
        data = []
        data_avg = []
        for i, model in enumerate(models):
            path = root + model.__class__.__name__
            data = []
            data_avg = []
            for project in projects:
                ast_f, ast_l = get_data_ast(project)           # ast
                tr_f, tr_l = get_data_tradition_10(project)    # tr
                # tr
                result, result_avg = split_data_train(tr_f, tr_l, model)
                data.append([project, '-', '-'])
                data = data + result
                data_avg.append([model.__class__.__name__, '-', '-'])
                data_avg.append([project, '-', '-'])
                data_avg.append(['tr'] + result_avg)
                write_excel_xls_hypoth(path + ".xlsx", project + "-tr", data)
                data = []

                # ast
                result, result_avg = split_data_train(ast_f, ast_l, model)   # 交叉验证
                data.append([project, '-', '-'])
                data = data + result
                data_avg.append(['ast-whole'] + result_avg)
                write_excel_xls_hypoth(path + ".xlsx", project + "-ast-whole", data)
                data = []
            else:
                write_excel_xls(root + model.__class__.__name__ + "_average.xls", 'sheet1', data_avg)

main()