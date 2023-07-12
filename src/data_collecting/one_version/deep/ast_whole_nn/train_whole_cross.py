import os
import torch
import copy
from torch.utils.data import DataLoader
import torch.utils.data as Data
import rnn as GRU
import torch.optim as optim
import torch.nn as nn
from torch.autograd import Variable
import pandas as pd
import numpy as np
import time
from sklearn.metrics import roc_auc_score
from sklearn.metrics import accuracy_score
from sklearn.metrics import precision_score
from sklearn.metrics import f1_score
from sklearn.metrics import recall_score
import openpyxl

def write_excel_xls_hypoth(path, sheet_name, value):
    if os.path.exists(path):
        data = openpyxl.load_workbook(path)
        table = data.create_sheet(sheet_name)
        table.title = sheet_name
        index = len(value)  
        for i in range(0, index):
            for j in range(0, len(value[i])):
                table.cell(row=1 + i, column=j + 1, value=str(value[i][j])) 
        data.save(path)  
    else:
        index = len(value)
        workbook = openpyxl.Workbook()  
        sheet = workbook.active  
        sheet.title = sheet_name 
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))  
        workbook.save(path) 
    print("xlsx格式表格写入数据成功！")

def adjust_learning_rate(optimizer, epoch):
    lr = learning_rate * (0.1 ** (epoch // 10))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr
    return optimizer

def get_data(project,h, k, type):  
    root = 'data_cross/' + project + '/' + project +'_' + str(h) +'/'+str(k)+'/' + type + '/features.pkl'
    data = pd.read_pickle(root)

    features = data['code'].values
    labels = data['label'].values

    labels = torch.from_numpy(labels)
    return features, labels


def handel_data(train_x):
    X = []
    train_x = train_x.tolist()
    n_x = len(train_x)
    max_lens = 0
    for i in range(n_x):
        length = len(train_x[i])
        if length>max_lens:
            max_lens = length

    for i in range(n_x):
        row = []
        for _ in range(max_lens - len(train_x[i])):
            row.append([0.0]*128)
        for vc in train_x[i]:
            row.append(vc)
        X.append(row)
    X= np.array(X)
    X = torch.from_numpy(X)
    
    return X, max_lens

os.environ['CUDA_VISIBLE_DEVICES'] = '0'

## parameter setting
epochs = 10
batch_size = 32
use_gpu = torch.cuda.is_available()
learning_rate = 0.01


projects = ['ambari'] 
data = []
data_avg = []
count = 30

for project in projects:
    print('*********************', project, '*********************')
    data.append([project, '-', '-'])
    data_avg.append([project, '-', '-'])
    acc_sum = 0
    auc_sum = 0
    f1_sum = 0
    pre_sum = 0
    recall_sum = 0
    time_epoch = 0
    for h in range(30):
        print('*********************', '第',str(h), '轮', '*********************')

        acc_cross_sum = 0
        auc_cross_sum = 0
        f1_cross_sum = 0
        pre_cross_sum = 0
        recall_cross_sum = 0
        # 10折交叉验证
        data_cross = []   # 保存交叉验证的中间结果
        for k in range(10):
            print('交叉验证----------', str(k))

            embedding_dim = 128
            hidden_dim = 128
            nlabel = 2

            model = GRU.GRUClassifier(embedding_dim=embedding_dim,hidden_dim=hidden_dim,label_size=nlabel, batch_size=batch_size, use_gpu=use_gpu)
            if use_gpu:
                model = model.cuda()

            train_x, train_y = get_data(project,h, k , 'train')
            train_x, train_max_lens = handel_data(train_x)
            train_data = Data.TensorDataset(train_x, train_y)
            train_loader = DataLoader(
                dataset=train_data,
                batch_size=batch_size,
                shuffle=True,
                num_workers=1,
            )
            test_x, test_y = get_data(project,h, k , 'test')
            test_x, max_lens = handel_data(test_x)
            test_data = Data.TensorDataset(test_x, test_y)
            test_loader = DataLoader(
                dataset=test_data,
                batch_size=batch_size,
                shuffle=True,
                num_workers=1,
            )
            dev_x, dev_y = get_data(project,h, k , 'dev')
            dev_x, max_lens = handel_data(dev_x)
            dev_data = Data.TensorDataset(dev_x, dev_y)
            dev_loader = DataLoader(
                dataset=dev_data,
                batch_size=batch_size,
                shuffle=True,
                num_workers=1,
            )

            optimizer = optim.Adam(model.parameters(), lr=learning_rate)
            loss_function = nn.CrossEntropyLoss()

            train_loss_ = []
            val_loss_ = []
            train_acc_ = []
            val_acc_ = []
            best_acc = 0.0
            best_loss =1.0 
            print('Start training...')
            # training procedure
            best_model = model
            for epoch in range(epochs):
                start_time = time.time()
                optimizer = adjust_learning_rate(optimizer, epoch)

                total_acc = 0.0
                total_loss = 0.0
                total = 0.0

                for iter, traindata in enumerate(train_loader):
                    train_inputs, train_labels = traindata
                
                    if use_gpu:
                        train_inputs, train_labels = Variable(train_inputs.cuda()), train_labels.cuda()
                    else: 
                        train_inputs = Variable(train_inputs)

                    model.zero_grad()
                    model.batch_size = len(train_labels)
                    model.hidden = model.init_hidden()
                    train_inputs = train_inputs.float()
                    output = model(train_inputs)
                    loss = loss_function(output, Variable(train_labels))
                    loss.backward()
                    optimizer.step()

                    # calc training acc
                    _, predicted = torch.max(output.data, 1)
                    total_acc += (predicted == train_labels).sum()
                    total += len(train_labels)
                    total_loss += loss.item()*len(train_inputs)

                train_loss_.append(total_loss / total)
                train_acc_.append(total_acc.item() / total)
                # validation epoch
                total_acc = 0.0
                total_loss = 0.0
                total = 0.0
                for iter, devdata in enumerate(dev_loader):
                    dev_inputs, dev_labels = devdata
                    if use_gpu:
                        dev_inputs, dev_labels = Variable(dev_inputs.cuda()), dev_labels.cuda()
                    else: 
                        dev_inputs = Variable(dev_inputs)

                    model.zero_grad()
                    model.batch_size = len(dev_labels)
                    model.hidden = model.init_hidden()
                    dev_inputs = dev_inputs.float()
                    output = model(dev_inputs)

                    loss = loss_function(output, Variable(dev_labels))

                # calc valing acc
                    _, predicted = torch.max(output.data, 1)
                    total_acc += (predicted == dev_labels).sum()
                    total += len(dev_labels)
                    total_loss += loss.item()*len(dev_inputs)
                    
                val_loss_.append(total_loss / total)
                val_acc_.append(total_acc.item() / total)
                end_time = time.time()
                if total_acc/total >= best_acc:
                    best_model = model
                print('[Epoch: %3d/%3d] Training Loss: %.4f, Validation Loss: %.4f,'
                ' Training Acc: %.3f, Validation Acc: %.3f, Time Cost: %.3f s'
                % (epoch + 1, epochs, train_loss_[epoch], val_loss_[epoch],
                 train_acc_[epoch], val_acc_[epoch], end_time - start_time))
                time_epoch = end_time - start_time


            ## testing epoch
            total_acc = 0.0
            total_loss = 0.0
            total = 0.0
            model = best_model
        
            print(project,'--------------------')
            for iter, testdata in enumerate(test_loader):
                test_inputs, test_labels = testdata
                if use_gpu:
                    test_inputs, test_labels = Variable(test_inputs.cuda()), test_labels.cuda()
                else: 
                    test_inputs = Variable(test_inputs)
                model.batch_size = len(test_labels)
                model.hidden = model.init_hidden()
                test_inputs = test_inputs.float()
                output = model(test_inputs)

                loss = loss_function(output, Variable(test_labels))

                _, predicted = torch.max(output.data, 1)

                total_acc += (predicted == test_labels).sum()
                total += len(test_labels)
                total_loss += loss.item() * len(test_inputs)
        
            print("Testing results(Acc):", total_acc.item() / total)
